"""Bulk import of expenses (as Invoices or Direct Expenses, optionally paid)
from an Excel workbook. Reuses expense_service/invoice_service exactly as the
single-record routers do, so a bulk-imported row produces identical
Expense/Invoice/Payment/AuditLog rows to one entered by hand - no parallel
creation logic.

Row-level validation happens against in-memory lookup maps (built once per
import) that resolve Project/Vendor/Category/SubCategory/Account by either
code or name (case-insensitive), since a human-authored spreadsheet won't
have internal ids.

Import is all-or-nothing: `process_workbook(..., dry_run=False)` only commits
if every row validates; otherwise nothing is written and the caller gets the
full list of per-row errors to fix and re-upload. `dry_run=True` runs the
same validation (and the same DB flushes, to catch numbering/FK issues) then
rolls back, for a preview pass from the frontend before the user confirms.
"""
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.models import Project, Vendor, ExpenseCategory, ExpenseSubCategory, Account
from app.models.enums import SourceType
from app.services import expense_service, invoice_service, project_scope_service

HEADERS = [
    "Row Type", "Date", "Project", "Vendor", "Category", "Sub Category", "Description",
    "Invoice Number", "Due Date", "Supplier Name", "Bill Number",
    "Base/Taxable Amount", "CGST", "SGST", "IGST", "Other Amount",
    "Pay Immediately", "Payment Date", "Account", "Payment Mode", "Reference Number",
]

ROW_TYPE_INVOICE = "INVOICE"
ROW_TYPE_DIRECT = "DIRECT"
_ROW_TYPE_SYNONYMS = {"INVOICE": ROW_TYPE_INVOICE, "DIRECT": ROW_TYPE_DIRECT, "EXPENSE": ROW_TYPE_DIRECT, "DIRECT_EXPENSE": ROW_TYPE_DIRECT}

_TRUE_VALUES = {"y", "yes", "true", "1"}


@dataclass
class RowResult:
    row_number: int
    row_type: str | None = None
    status: str = "OK"  # OK | ERROR
    errors: list[str] = field(default_factory=list)
    expense_number: str | None = None
    invoice_number: str | None = None
    payment_number: str | None = None
    total_amount: str | None = None


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _cell_date(v, errors: list[str], label: str):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _norm(v)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    errors.append(f"{label} '{s}' is not a valid date (use YYYY-MM-DD)")
    return None


def _cell_decimal(v, errors: list[str], label: str, default: Decimal = Decimal("0")) -> Decimal:
    if v is None or v == "":
        return default
    try:
        return Decimal(str(v))
    except InvalidOperation:
        errors.append(f"{label} '{v}' is not a valid number")
        return default


class _Lookups:
    def __init__(self, db: Session):
        self.projects = self._index(db.query(Project).all(), "code", "name")
        self.vendors = self._index(db.query(Vendor).all(), "vendor_code", "vendor_name")
        self.accounts = self._index(db.query(Account).all(), None, "account_name")
        self.categories = self._index(db.query(ExpenseCategory).all(), None, "name")
        sub_cats = db.query(ExpenseSubCategory).all()
        self.sub_categories_by_category: dict[int, dict[str, ExpenseSubCategory]] = {}
        for sc in sub_cats:
            self.sub_categories_by_category.setdefault(sc.category_id, {})[sc.name.strip().lower()] = sc

    @staticmethod
    def _index(rows, code_attr, name_attr):
        idx = {}
        for r in rows:
            if code_attr:
                idx[getattr(r, code_attr).strip().lower()] = r
            idx[getattr(r, name_attr).strip().lower()] = r
        return idx

    def find(self, idx: dict, value: str):
        return idx.get(value.strip().lower()) if value else None


def _read_rows(file_bytes: bytes) -> list[dict]:
    import io
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_norm(h).lower() for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if raw is None or all(c is None or _norm(c) == "" for c in raw):
            continue
        out.append({header[i]: raw[i] for i in range(min(len(header), len(raw)))})
    return out


def process_workbook(db: Session, file_bytes: bytes, user, dry_run: bool) -> dict:
    lookups = _Lookups(db)
    accounts_scoped = user.role.name == "ACCOUNTS"
    raw_rows = _read_rows(file_bytes)

    results: list[RowResult] = []
    any_error = False

    for i, raw in enumerate(raw_rows, start=2):  # +1 for header, +1 for 1-index
        r = RowResult(row_number=i)
        errors = r.errors

        row_type_raw = _norm(raw.get("row type")).upper()
        row_type = _ROW_TYPE_SYNONYMS.get(row_type_raw)
        if not row_type:
            errors.append(f"Row Type '{row_type_raw}' must be INVOICE or DIRECT")
            r.status = "ERROR"
            results.append(r)
            any_error = True
            continue
        r.row_type = row_type

        tx_date = _cell_date(raw.get("date"), errors, "Date")

        project = None
        project_name = _norm(raw.get("project"))
        if project_name:
            project = lookups.find(lookups.projects, project_name)
            if not project:
                errors.append(f"Project '{project_name}' not found")

        category = None
        category_name = _norm(raw.get("category"))
        if not category_name:
            errors.append("Category is required")
        else:
            category = lookups.find(lookups.categories, category_name)
            if not category:
                errors.append(f"Category '{category_name}' not found")

        sub_category = None
        sub_category_name = _norm(raw.get("sub category"))
        if sub_category_name and category:
            sub_category = lookups.sub_categories_by_category.get(category.id, {}).get(sub_category_name.lower())
            if not sub_category:
                errors.append(f"Sub Category '{sub_category_name}' not found under '{category_name}'")

        description = _norm(raw.get("description")) or None
        base_amount = _cell_decimal(raw.get("base/taxable amount"), errors, "Base/Taxable Amount")
        cgst = _cell_decimal(raw.get("cgst"), errors, "CGST")
        sgst = _cell_decimal(raw.get("sgst"), errors, "SGST")
        igst = _cell_decimal(raw.get("igst"), errors, "IGST")
        other_amount = _cell_decimal(raw.get("other amount"), errors, "Other Amount")

        pay_immediately = _norm(raw.get("pay immediately")).lower() in _TRUE_VALUES
        payment_date = _cell_date(raw.get("payment date"), errors, "Payment Date")
        account = None
        account_name = _norm(raw.get("account"))
        payment_mode = _norm(raw.get("payment mode")) or None
        reference_number = _norm(raw.get("reference number")) or None
        if pay_immediately:
            if not payment_date:
                errors.append("Payment Date is required when Pay Immediately is set")
            if not payment_mode:
                errors.append("Payment Mode is required when Pay Immediately is set")
            if not account_name:
                errors.append("Account is required when Pay Immediately is set")
            else:
                account = lookups.find(lookups.accounts, account_name)
                if not account:
                    errors.append(f"Account '{account_name}' not found")

        if accounts_scoped:
            try:
                project_scope_service.assert_project_in_scope(db, user, project.id if project else None)
            except Exception:
                errors.append("You are not assigned to this row's project")

        vendor = None
        invoice_number = None
        due_date = None
        if row_type == ROW_TYPE_INVOICE:
            vendor_name = _norm(raw.get("vendor"))
            if not vendor_name:
                errors.append("Vendor is required for an INVOICE row")
            else:
                vendor = lookups.find(lookups.vendors, vendor_name)
                if not vendor:
                    errors.append(f"Vendor '{vendor_name}' not found")
            invoice_number = _norm(raw.get("invoice number")) or None
            if not invoice_number:
                errors.append("Invoice Number is required for an INVOICE row")
            due_date = _cell_date(raw.get("due date"), errors, "Due Date")
            if not tx_date:
                errors.append("Date (invoice date) is required")

        supplier_name = None
        bill_number = None
        if row_type == ROW_TYPE_DIRECT:
            supplier_name = _norm(raw.get("supplier name")) or None
            bill_number = _norm(raw.get("bill number")) or None
            if not tx_date:
                errors.append("Date (expense date) is required")
            gst_amount = cgst + sgst + igst

        total = base_amount + cgst + sgst + igst + other_amount if row_type == ROW_TYPE_INVOICE else base_amount + (cgst + sgst + igst) + other_amount
        if total <= 0:
            errors.append("Row total amount must be greater than zero")

        if errors:
            r.status = "ERROR"
            results.append(r)
            any_error = True
            continue

        try:
            if row_type == ROW_TYPE_INVOICE:
                invoice = invoice_service.create_invoice(
                    db, invoice_number=invoice_number, vendor_id=vendor.id, invoice_date=tx_date, due_date=due_date,
                    project_id=project.id if project else None, description=description, taxable_amount=base_amount,
                    cgst=cgst, sgst=sgst, igst=igst, other_tax=other_amount, category_id=category.id,
                    sub_category_id=sub_category.id if sub_category else None, created_by=user.id,
                    pay_immediately=pay_immediately, payment_date=payment_date,
                    account_id=account.id if account else None, payment_mode=payment_mode,
                    reference_number=reference_number, remarks=None,
                )
                r.invoice_number = invoice.invoice_number
                r.expense_number = invoice.expense.expense_number
                r.total_amount = str(invoice.total_amount)
                if pay_immediately:
                    r.payment_number = invoice.expense.payments[-1].payment_number if invoice.expense.payments else None
            else:
                expense = expense_service.create_expense_record(
                    db, source_type=SourceType.DIRECT_EXPENSE, source_id=None, expense_date=tx_date,
                    project_id=project.id if project else None, vendor_id=None, employee_id=None,
                    category_id=category.id, sub_category_id=sub_category.id if sub_category else None,
                    description=description, base_amount=base_amount, gst_amount=gst_amount,
                    other_amount=other_amount, created_by=user.id, supplier_name=supplier_name, bill_number=bill_number,
                )
                r.expense_number = expense.expense_number
                r.total_amount = str(expense.total_amount)
                if pay_immediately:
                    payment = expense_service.pay_expense_immediately(
                        db, expense=expense, payment_date=payment_date, account_id=account.id,
                        payment_mode=payment_mode, reference_number=reference_number, remarks=None, created_by=user.id,
                    )
                    r.payment_number = payment.payment_number
            db.flush()
        except Exception as exc:  # noqa: BLE001 - surface any service-layer validation as a row error
            r.status = "ERROR"
            r.errors.append(str(getattr(exc, "detail", exc)))
            any_error = True

        results.append(r)

    summary = {
        "total_rows": len(results),
        "ok_rows": sum(1 for r in results if r.status == "OK"),
        "error_rows": sum(1 for r in results if r.status == "ERROR"),
        "committed": False,
    }

    if dry_run or any_error:
        db.rollback()
    else:
        db.commit()
        summary["committed"] = True

    summary["rows"] = [
        {
            "row_number": r.row_number, "row_type": r.row_type, "status": r.status, "errors": r.errors,
            "expense_number": r.expense_number, "invoice_number": r.invoice_number,
            "payment_number": r.payment_number, "total_amount": r.total_amount,
        }
        for r in results
    ]
    return summary


def build_template_workbook() -> bytes:
    import io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    ws.append(HEADERS)
    ws.append([
        "INVOICE", "2026-04-01", "GEN", "Acme Supplies", "Office Supplies", "", "Stationery for April",
        "INV-1001", "2026-05-01", "", "", 1000, 90, 90, 0, 0, "N", "", "", "", "",
    ])
    ws.append([
        "DIRECT", "2026-04-02", "GEN", "", "Travel", "", "Taxi fare", "", "", "Local Taxi", "BILL-55",
        500, 0, 0, 0, 0, "Y", "2026-04-02", "Main Bank Account", "CASH", "",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
